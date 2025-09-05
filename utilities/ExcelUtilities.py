import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[file]
    return sheet.max_column

def readData(file,sheetName,rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,colnum).value

def writeData(file, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,colnum).value = data
    workbook.save(file)

def fillGreenColor(file,sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenfill = PatternFill(start_color= '008000',
                            end_color= '008000',
                            fill_type= 'solid')
    sheet.cell(rownum,colnum).fill = greenfill
    workbook.save(file)

def fillRedColor(file, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='FB1604',
                          end_color= 'FB1604',
                          fill_type= 'solid')
    sheet.cell(rownum,colnum).fill = redFill
    workbook.save(file)

